#fetch the data from the excel file using python module openpyxl
#importing module
import openpyxl
#give location of the file
path = 'data9.xlsx'
#creating an object
data9_object = openpyxl.load_workbook(path)
#storing the data from active sheet into an object called data9
data9_object=data9_object.active
data9_sheet = data9_object.title
get_data = data9_object.cell(row=1,column=1)
data9_object.cell(row=1,column=1).value = 'changed'
print(get_data.value)


#print the total number of rows and columns
print('column = ',data9_object.max_column, 'row = ',data9_object.max_row)
print(data9_sheet)
#find values from excel
#changing cells and values

